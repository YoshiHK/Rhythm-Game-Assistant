from __future__ import annotations
"""[qa_reporter.py](https://onedrive.live.com?cid=D5D62A1EF303BA22&id=D5D62A1EF303BA22!s38b47db446c649b3bf7f7c91060fdb2a&EntityRepresentationId=4d061505-d8aa-4ca8-9970-c7475518461f)
Aggregated QA reporting utilities for Phase 3 (UMI).

Control-plane only:
- Collects run diagnostics and formats them for humans/CI artifacts.
- Does NOT gate execution and does NOT alter gameplay semantics.
"""

from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional
import json


@dataclass
class QASummary:
    """Aggregated QA summary for an ingestion run.

    Fields
    ------
    total_charts   : total number of charts processed
    total_success  : number of charts that passed (no fatal error)
    total_failed   : number of charts that raised an error
    by_game        : { game_id: { "success": int, "failed": int } }
    failures       : list of { "game_id", "song_id", "error" }
    metadata_stats : aggregated metadata parity stats (diagnostic counters)
                     e.g. combo_mismatch_charts, bpm_mismatch_charts, duration_mismatch_charts
    """

    total_charts: int
    total_success: int
    total_failed: int
    by_game: Dict[str, Dict[str, int]]
    failures: List[Dict[str, Any]] = field(default_factory=list)
    metadata_stats: Dict[str, Any] = field(default_factory=dict)

    # -----------------------------
    # Pretty text
    # -----------------------------
    def to_pretty_string(self) -> str:
        lines: List[str] = []
        lines.append("=== Unified Ingestion QA Summary ===")
        lines.append(f"Total Charts: {self.total_charts}")
        lines.append(f"Success:      {self.total_success}")
        lines.append(f"Failed:       {self.total_failed}")
        lines.append("")
        lines.append("By Game:")
        for game, stats in self.by_game.items():
            lines.append(f"  - {game}: success={stats.get('success', 0)}, failed={stats.get('failed', 0)}")

        if self.failures:
            lines.append("\n--- Failures ---")
            for fail in self.failures:
                lines.append(f"[{fail.get('game_id')}] {fail.get('song_id')}: {fail.get('error')}")

        if self.metadata_stats:
            lines.append("\n--- Metadata Stats (from diagnostics) ---")
            for key, value in self.metadata_stats.items():
                lines.append(f"{key}: {value}")

        return "\n".join(lines)

    # -----------------------------
    # JSON Export
    # -----------------------------
    def to_json_dict(self) -> Dict[str, Any]:
        return {
            "total_charts": self.total_charts,
            "total_success": self.total_success,
            "total_failed": self.total_failed,
            "by_game": self.by_game,
            "failures": self.failures,
            "metadata_stats": self.metadata_stats,
        }

    def to_json_string(self, indent: int = 2) -> str:
        return json.dumps(self.to_json_dict(), indent=indent, ensure_ascii=False)

    def save_json(self, path: str, indent: int = 2) -> None:
        """Save the QA summary as a JSON file on disk."""
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.to_json_dict(), f, indent=indent, ensure_ascii=False)

    # -----------------------------
    # QA Index & Pivots
    # -----------------------------
    def compute_qa_index(self) -> float:
        """Compute a simple QA Index Score in [0, 100].

        Base = 100
        -2 per failed chart
        -1 per mismatch count (combo / bpm / duration)
        Floored at 0.
        """
        score = 100
        score -= 2 * int(self.total_failed)
        score -= int(self.metadata_stats.get("combo_mismatch_charts", 0) or 0)
        score -= int(self.metadata_stats.get("bpm_mismatch_charts", 0) or 0)
        score -= int(self.metadata_stats.get("duration_mismatch_charts", 0) or 0)
        return max(0.0, float(score))

    def build_pivot_game(self) -> List[Dict[str, Any]]:
        pivot: List[Dict[str, Any]] = []
        for game_id, stats in self.by_game.items():
            s = int(stats.get("success", 0) or 0)
            f = int(stats.get("failed", 0) or 0)
            total = s + f
            rate = (s / total * 100.0) if total > 0 else 0.0
            pivot.append({"game_id": game_id, "success": s, "failed": f, "success_rate": rate})
        return pivot

    def build_pivot_errors(self) -> List[Dict[str, Any]]:
        counter: Dict[str, int] = {}
        for fail in self.failures:
            msg = str(fail.get("error", "Unknown Error"))
            counter[msg] = counter.get(msg, 0) + 1
        return [{"error": k, "count": v} for k, v in counter.items()]

    # -----------------------------
    # Excel Export (optional)
    # -----------------------------
    def save_to_excel(self, db_path: str, sheet_name: str = "QA_Report", overwrite: bool = True) -> None:
        """Write the QA summary into an Excel workbook as a styled QA sheet.

        Requires openpyxl.
        """
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.chart import BarChart, Reference
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule

        wb = load_workbook(db_path)
        if sheet_name in wb.sheetnames:
            if overwrite:
                ws_old = wb[sheet_name]
                wb.remove(ws_old)
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)

        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def style_section(cell):
            cell.font = Font(bold=True, size=13)
            cell.fill = PatternFill("solid", fgColor="B7E1CD")
            cell.border = thin
            cell.alignment = Alignment(horizontal="left")

        def style_header(row):
            for c in row:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="DDDDDD")
                c.border = thin
                c.alignment = Alignment(horizontal="center")

        def zebra(row, odd: bool):
            fill = "FFFFFF" if odd else "F5F9FF"
            for c in row:
                c.fill = PatternFill("solid", fgColor=fill)
                c.border = thin
                c.alignment = Alignment(horizontal="left")

        def heat_color(i: int) -> str:
            if i <= 0:
                return "FFFFFF"
            if i == 1:
                return "FFEEEE"
            if i == 2:
                return "FFDDDD"
            if i == 3:
                return "FFCCCC"
            return "FFBBBB"

        # 1) Top section
        ws.append(["Unified Ingestion QA Summary"])
        style_section(ws["A1"])
        ws.append([])
        ws.append(["Metric", "Value"])
        style_header(ws[3])
        ws.append(["total_charts", self.total_charts])
        ws.append(["total_success", self.total_success])
        ws.append(["total_failed", self.total_failed])

        # 2) Metadata stats
        ws.append([])
        ws.append(["Metadata Stats (from diagnostics)"])
        style_section(ws[f"A{ws.max_row}"])
        ws.append(["Key", "Value"])
        meta_header_row = ws.max_row
        style_header(ws[meta_header_row])
        meta_values_start = meta_header_row + 1
        for k, v in self.metadata_stats.items():
            ws.append([k, v])
        meta_values_end = ws.max_row

        for row in ws.iter_rows(min_row=meta_values_start, max_row=meta_values_end):
            key = row[0].value
            val = row[1].value
            if key in ["combo_mismatch_charts", "bpm_mismatch_charts", "duration_mismatch_charts"] and isinstance(val, int):
                row[1].fill = PatternFill("solid", fgColor=heat_color(val))
                row[1].border = thin

        # 3) By-game section
        ws.append([])
        ws.append(["By Game"])
        style_section(ws[f"A{ws.max_row}"])
        ws.append(["game_id", "success", "failed"])
        by_game_header_row = ws.max_row
        style_header(ws[by_game_header_row])
        start_by_game = by_game_header_row + 1
        zebra_counter = 0
        for game_id, stats in self.by_game.items():
            zebra_counter += 1
            ws.append([game_id, stats.get("success", 0), stats.get("failed", 0)])
            zebra(ws[ws.max_row], zebra_counter % 2 == 1)
        end_by_game = ws.max_row

        # 4) Failures section
        if self.failures:
            ws.append([])
            ws.append(["Failures"])
            style_section(ws[f"A{ws.max_row}"])
            ws.append(["game_id", "song_id", "error"])
            failures_header_row = ws.max_row
            style_header(ws[failures_header_row])
            zebra_counter = 0
            for fail in self.failures:
                zebra_counter += 1
                ws.append([fail.get("game_id"), fail.get("song_id"), fail.get("error")])
                zebra(ws[ws.max_row], zebra_counter % 2 == 1)

        # 5) QA Index Score
        qa_score = self.compute_qa_index()
        ws.append([])
        ws.append(["QA Index Score", qa_score])
        score_row = ws.max_row
        ws[f"A{score_row}"].font = Font(bold=True)
        ws[f"B{score_row}"].font = Font(bold=True)

        # 6) Pivot: per-game
        ws.append([])
        ws.append(["Pivot: Per Game Summary"])
        style_section(ws[f"A{ws.max_row}"])
        ws.append(["game_id", "success", "failed", "success_rate%"])
        pivot_game_header_row = ws.max_row
        style_header(ws[pivot_game_header_row])
        pivot_game_start = pivot_game_header_row + 1
        zebra_counter = 0
        for item in self.build_pivot_game():
            zebra_counter += 1
            ws.append([item["game_id"], item["success"], item["failed"], item["success_rate"]])
            zebra(ws[ws.max_row], zebra_counter % 2 == 1)
        pivot_game_end = ws.max_row

        # 7) Pivot: error types
        pivot_errors = self.build_pivot_errors()
        pivot_errors_start: Optional[int] = None
        pivot_errors_end: Optional[int] = None
        if pivot_errors:
            ws.append([])
            ws.append(["Pivot: Failure Error Types"])
            style_section(ws[f"A{ws.max_row}"])
            ws.append(["error", "count"])
            pivot_err_header_row = ws.max_row
            style_header(ws[pivot_err_header_row])
            pivot_errors_start = pivot_err_header_row + 1
            zebra_counter = 0
            for item in pivot_errors:
                zebra_counter += 1
                ws.append([item["error"], item["count"]])
                zebra(ws[ws.max_row], zebra_counter % 2 == 1)
            pivot_errors_end = ws.max_row

        # 8) Charts
        if end_by_game >= start_by_game:
            chart = BarChart()
            chart.title = "Success vs Fail (By Game)"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Game"
            cats = Reference(ws, min_col=1, min_row=start_by_game, max_row=end_by_game)
            success_ref = Reference(ws, min_col=2, min_row=by_game_header_row, max_row=end_by_game)
            fail_ref = Reference(ws, min_col=3, min_row=by_game_header_row, max_row=end_by_game)
            chart.add_data(success_ref, titles_from_data=True)
            chart.add_data(fail_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f"J{start_by_game}")

        if pivot_game_end >= pivot_game_start:
            chart2 = BarChart()
            chart2.title = "Success vs Fail (Pivot)"
            chart2.y_axis.title = "Count"
            chart2.x_axis.title = "Game"
            cats2 = Reference(ws, min_col=1, min_row=pivot_game_start, max_row=pivot_game_end)
            su_ref = Reference(ws, min_col=2, min_row=pivot_game_header_row, max_row=pivot_game_end)
            fa_ref = Reference(ws, min_col=3, min_row=pivot_game_header_row, max_row=pivot_game_end)
            chart2.add_data(su_ref, titles_from_data=True)
            chart2.add_data(fa_ref, titles_from_data=True)
            chart2.set_categories(cats2)
            ws.add_chart(chart2, f"J{pivot_game_start}")

        # 9) Conditional formatting
        if meta_values_end >= meta_values_start:
            meta_range = f"B{meta_values_start}:B{meta_values_end}"
            color_scale = ColorScaleRule(
                start_type="min", start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFF3CD",
                end_type="max", end_color="FF9999",
            )
            ws.conditional_formatting.add(meta_range, color_scale)

        if end_by_game >= start_by_game:
            ws.conditional_formatting.add(
                f"B{start_by_game}:B{end_by_game}",
                DataBarRule(color="90EE90", showValue="None"),
            )
            ws.conditional_formatting.add(
                f"C{start_by_game}:C{end_by_game}",
                DataBarRule(color="FF9999", showValue="None"),
            )

        if pivot_game_end >= pivot_game_start:
            rate_range = f"D{pivot_game_start}:D{pivot_game_end}"
            icon_rule = IconSetRule(icon_style="3Arrows", type="percent", values=[0, 50, 80], showValue="None")
            ws.conditional_formatting.add(rate_range, icon_rule)

        if pivot_errors_start is not None and pivot_errors_end is not None and pivot_errors_end >= pivot_errors_start:
            ws.conditional_formatting.add(
                f"B{pivot_errors_start}:B{pivot_errors_end}",
                DataBarRule(color="FFA07A", showValue="None"),
            )

        # 10) Freeze panes
        ws.freeze_panes = "A3"

        # 11) Auto-fit columns
        for col in ws.columns:
            max_len = 0
            letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        wb.save(db_path)


__all__ = ["QASummary"]