"""Definitive Excel writer for UMI Phase 3.

Goals:
- Preserve the public surface used by the current pipeline: ExcelWriter(...),
  insert_row(...), save().
- Make db_path optional so orchestrator/batch flows do not fail when the CLI
  omits --db-path.
- Scale across all current games defined in config/games.json and remain safe
  for future additions by deriving sheet names dynamically and creating sheets
  lazily when needed.
- Accept both rich canonical rows and minimal/fallback rows emitted by early
  adapters/validators during Phase 3 evolution.

This writer upgrades persistence wiring only. It does not introduce gameplay
semantics or modify completed analysis phases.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional

from openpyxl import Workbook, load_workbook


DEFAULT_DB_FILENAME = "Song Database (full).xlsx"

# Keep the existing canonical column order so downstream workbook consumers do
# not need to change when the writer is upgraded.
CANONICAL_COLUMNS: List[str] = [
    "song_id",
    "name",
    "tier",
    "level",
    "difficulty_label",
    "difficulty_code",
    "rating_raw",
    "note_total_db",
    "note_total_chart",
    "note_delta",
    "chart_path",
    "duration_ms",
    "bpm",
]

# Alias map allows minimal Phase-3 rows (for example chart_id/title/difficulty)
# to be normalized into the canonical Excel layout without failing the writer.
FIELD_ALIASES: Dict[str, Iterable[str]] = {
    "song_id": ("song_id", "chart_id"),
    "name": ("name", "title"),
    "tier": ("tier",),
    "level": ("level",),
    "difficulty_label": ("difficulty_label", "difficulty"),
    "difficulty_code": ("difficulty_code", "difficulty"),
    "rating_raw": ("rating_raw",),
    "note_total_db": ("note_total_db",),
    "note_total_chart": ("note_total_chart",),
    "note_delta": ("note_delta",),
    "chart_path": ("chart_path", "source_file"),
    "duration_ms": ("duration_ms",),
    "bpm": ("bpm",),
}

# Compatibility overrides matching the legacy writer sheet names.
BUILTIN_SHEET_NAME_OVERRIDES: Dict[str, str] = {
    "arcaea": "Arcaea",
    "proseka": "Project SEKAI",
    "bandori": "BanG Dream",
}


class ExcelWriter:
    def __init__(
        self,
        db_path: Optional[str | Path] = None,
        *,
        games_config_path: Optional[str | Path] = None,
        auto_create: bool = True,
        create_missing_sheets: bool = True,
        precreate_known_sheets: bool = False,
    ) -> None:
        self.db_path = Path(db_path) if db_path else Path.cwd() / DEFAULT_DB_FILENAME
        self.games_config_path = (
            Path(games_config_path)
            if games_config_path
            else self._default_games_config_path()
        )
        self.auto_create = bool(auto_create)
        self.create_missing_sheets = bool(create_missing_sheets)
        self.precreate_known_sheets = bool(precreate_known_sheets)
        self.sheet_names = self._load_sheet_names()
        self.wb = self._load_or_create_workbook()
        self._ensure_header_rows()
        if self.precreate_known_sheets:
            self._ensure_known_sheets()

    def _default_games_config_path(self) -> Path:
        # writers/excel_writer.py -> ../config/games.json
        return Path(__file__).resolve().parents[1] / "config" / "games.json"

    def _load_sheet_names(self) -> Dict[str, str]:
        sheet_names = dict(BUILTIN_SHEET_NAME_OVERRIDES)
        if self.games_config_path.exists():
            try:
                data = json.loads(self.games_config_path.read_text(encoding="utf-8"))
                for item in data.get("games", []):
                    game_id = item.get("game_id")
                    display_name = item.get("display_name")
                    if not game_id or not display_name:
                        continue
                    sheet_names.setdefault(str(game_id), str(display_name))
            except Exception:
                # Keep the builtin mapping as the safe fallback.
                pass
        return sheet_names

    def _load_or_create_workbook(self):
        if self.db_path.exists():
            return load_workbook(self.db_path)
        if not self.auto_create:
            raise FileNotFoundError(f"Workbook not found: {self.db_path}")
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        # Keep a single seed worksheet so a brand-new workbook stays valid.
        ws = wb.active
        ws.title = "_meta"
        ws.append(["created_by", "ExcelWriter"])
        return wb

    def _ensure_header_rows(self) -> None:
        # Seed any already-existing game sheets that are empty/malformed.
        for name in self.wb.sheetnames:
            if name == "_meta":
                continue
            ws = self.wb[name]
            if ws.max_row == 1 and ws.max_column == 1 and ws[1][0].value is None:
                ws.delete_rows(1, 1)
                ws.append(CANONICAL_COLUMNS)
            elif ws.max_row == 0:
                ws.append(CANONICAL_COLUMNS)

    def _ensure_known_sheets(self) -> None:
        for game_id in self.sheet_names:
            self._ensure_sheet(game_id)

    def _resolve_sheet_name(self, game_id: str) -> str:
        game_id = str(game_id)
        if game_id in self.sheet_names:
            return self.sheet_names[game_id]
        # Future-safe fallback: use the raw game_id and remember it.
        self.sheet_names[game_id] = game_id
        return game_id

    def _ensure_sheet(self, game_id: str):
        sheet_name = self._resolve_sheet_name(game_id)
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            # If the sheet exists but lacks the canonical header, add it.
            if ws.max_row == 1 and ws.max_column == 1 and ws[1][0].value is None:
                ws.delete_rows(1, 1)
                ws.append(CANONICAL_COLUMNS)
            elif ws.max_row == 0:
                ws.append(CANONICAL_COLUMNS)
            elif ws.max_row == 1:
                header = [cell.value for cell in ws[1]]
                if header != CANONICAL_COLUMNS:
                    # Do not destroy existing content; create the canonical header only
                    # when the sheet is empty. Existing divergent sheets are preserved.
                    pass
            return ws

        if not self.create_missing_sheets:
            raise KeyError(f"Missing worksheet for game_id='{game_id}'")

        ws = self.wb.create_sheet(title=sheet_name)
        ws.append(CANONICAL_COLUMNS)
        return ws

    def _first_present_value(self, row: Mapping[str, Any], keys: Iterable[str]) -> Any:
        for key in keys:
            if key in row:
                value = row.get(key)
                if value not in (None, ""):
                    return value
        return None

    def _normalize_row(self, canonical_row: Mapping[str, Any]) -> Dict[str, Any]:
        normalized: Dict[str, Any] = {}
        for output_key, aliases in FIELD_ALIASES.items():
            normalized[output_key] = self._first_present_value(canonical_row, aliases)
        return normalized

    def _dedupe_key(self, normalized_row: Mapping[str, Any]) -> tuple[Any, Any]:
        return (
            normalized_row.get("song_id"),
            normalized_row.get("difficulty_code"),
        )

    def _header_index(self, sheet) -> Dict[str, int]:
        if sheet.max_row < 1:
            return {}
        return {cell.value: idx for idx, cell in enumerate(sheet[1]) if cell.value is not None}

    def _find_existing_row_index(self, sheet, dedupe_key: tuple[Any, Any]) -> Optional[int]:
        song_id, difficulty_code = dedupe_key
        if not song_id:
            return None

        header_index = self._header_index(sheet)
        song_idx = header_index.get("song_id")
        diff_idx = header_index.get("difficulty_code")
        if song_idx is None:
            return None

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                continue

            row_song_id = row[song_idx] if song_idx < len(row) else None
            row_difficulty = row[diff_idx] if (diff_idx is not None and diff_idx < len(row)) else None

            if difficulty_code is None:
                if row_song_id == song_id:
                    return row_idx
            else:
                if (row_song_id, row_difficulty) == (song_id, difficulty_code):
                    return row_idx

        return None

    # ------------------------------------------------------------
    # New separated row operations
    # ------------------------------------------------------------

    def _is_empty_cell_value(self, value: Any) -> bool:
        """
        Safe-update rule:
        treat None / empty string as missing.
        Do NOT treat 0 as empty, because note counts / bpm / duration may legitimately be 0.
        """
        return value is None or value == ""

    def lookup_row(
        self,
        game_id: str,
        canonical_row: Mapping[str, Any],
    ) -> tuple[Any, Dict[str, Any], tuple[Any, Any], Optional[int]]:
        """
        Normalize the row, resolve its dedupe key, and look up an existing row.

        Returns:
            (sheet, normalized_row, dedupe_key, existing_row_idx)
        """
        sheet = self._ensure_sheet(game_id)
        normalized = self._normalize_row(canonical_row)
        dedupe_key = self._dedupe_key(normalized)
        existing_row_idx = self._find_existing_row_index(sheet, dedupe_key)
        return sheet, normalized, dedupe_key, existing_row_idx

    def update_row(
        self,
        game_id: str,
        canonical_row: Mapping[str, Any],
        row_idx: int,
        *,
        overwrite_existing: bool = False,
        overwrite_columns: Optional[Iterable[str]] = None,
    ) -> bool:
        """
        Safe-update policy:
        - by default, only fill missing existing cells
        - if overwrite_existing=True, overwrite all columns
        - if overwrite_columns is provided, those columns are always overwritten

        Returns:
            False (updated existing row, matching old upsert semantics)
        """
        sheet = self._ensure_sheet(game_id)
        normalized = self._normalize_row(canonical_row)
        overwrite_columns_set = set(overwrite_columns or ())

        for col_idx, col_name in enumerate(CANONICAL_COLUMNS, start=1):
            new_value = normalized.get(col_name)
            cell = sheet.cell(row=row_idx, column=col_idx)

            should_overwrite = (
                overwrite_existing
                or col_name in overwrite_columns_set
                or self._is_empty_cell_value(cell.value)
            )

            if should_overwrite and new_value not in (None, ""):
                cell.value = new_value

        return False

    def insert_row(self, game_id: str, canonical_row: Mapping[str, Any]) -> bool:
        """
        Insert a new row only.

        Returns:
            True if a new row was appended.
        """
        sheet = self._ensure_sheet(game_id)
        normalized = self._normalize_row(canonical_row)
        ordered = [normalized.get(col) for col in CANONICAL_COLUMNS]
        sheet.append(ordered)
        return True

    def upsert_row(
        self,
        game_id: str,
        canonical_row: Mapping[str, Any],
        *,
        overwrite_existing: bool = False,
        overwrite_columns: Optional[Iterable[str]] = None,
    ) -> bool:
        """
        Preserve previous write_rows behavior, but with explicit lookup/update/insert flow.

        Returns:
            True if inserted
            False if updated
        """
        sheet, normalized, dedupe_key, existing_row_idx = self.lookup_row(game_id, canonical_row)

        if existing_row_idx is not None:
            return self.update_row(
                game_id,
                normalized,
                existing_row_idx,
                overwrite_existing=overwrite_existing,
                overwrite_columns=overwrite_columns,
            )

        return self.insert_row(game_id, normalized)

    def write_rows(self, rows, db_path=None):
        # ✅ honor db_path from orchestrator / CLI
        if db_path:
            new_path = Path(db_path)
            if new_path != self.db_path:
                self.db_path = new_path
                self.wb = self._load_or_create_workbook()
                self._ensure_header_rows()
                if self.precreate_known_sheets:
                    self._ensure_known_sheets()

        print("WRITE_ROWS RECEIVED:", len(rows))

        for item in rows:
            if isinstance(item, dict):
                game_id = item.get("game_id") or item.get("game")

                if "canonical_row" in item:
                    canonical_row = item["canonical_row"]
                else:
                    canonical_row = item
                    canonical_row = {
                        "song_id": canonical_row.get("chart_id"),
                        "name": canonical_row.get("title"),
                        "difficulty_label": canonical_row.get("difficulty"),
                        "difficulty_code": canonical_row.get("difficulty"),
                        "chart_path": canonical_row.get("source_file"),
                    }

                self.upsert_row(game_id, canonical_row)

        self.save()
       
    def save(self) -> None:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        print("DB SAVED TO:", self.db_path)
        self.wb.save(self.db_path)


