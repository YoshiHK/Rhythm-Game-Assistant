#!/usr/bin/env python3
from __future__ import annotations

"""
adapter_bandori.py

Bandori adapter (fallback-capable version) for UMI.

Purpose of this implementation:
- keep the adapter import-safe when legacy bandori_model / bandori_chart modules are absent
- support routing-level acceptance for .json plus baseline fallback .html/.mht
- emit a minimal but valid canonical payload so UMI can produce rows
- preserve a future path for richer Bestdori/HTML parsing without blocking routing today

IMPORTANT:
- .json remains the primary machine-readable source format.
- .html / .mht support is heuristic and filename/text based.
- This adapter intentionally avoids gameplay semantics, tips generation, and heavy inference.
"""

import json
import re
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from ..base_adapter import BaseAdapter
from ..adapter_payload_utils import (
    build_adapter_metadata,
    attach_adapter_metadata,
    attach_internal_metadata,
)
from ..common_adapter_utils import (
    with_baseline_fallback_extensions,
)

DEFAULT_SONG_DB_FILENAME = "Song Database (full).xlsx"
DEFAULT_BANDORI_SHEET_NAME = "BanG Dream"

_BESTDORI_HINT_TOKENS = {
    "bestdori",
    "bang dream",
    "gbp resource site",
    "bandori",
}

_DIFFICULTY_TOKENS = [
    "special",
    "expert",
    "hard",
    "normal",
    "easy",
]


def _looks_like_bandori_export(path: Path, text: Optional[str] = None) -> bool:
    name = path.name.casefold()
    if any(tok in name for tok in _BESTDORI_HINT_TOKENS):
        return True
    if text:
        lowered = text.casefold()
        return any(tok in lowered for tok in _BESTDORI_HINT_TOKENS)
    return False


def _infer_difficulty_from_name(path: Path) -> Optional[str]:
    name = path.stem.casefold()
    for token in _DIFFICULTY_TOKENS:
        if f"[{token}]" in name or token in name:
            return token.upper()
    return None


def _infer_chart_id_from_path(path: Path) -> str:
    return path.stem.strip() or path.name


def _infer_title_from_bestdori_name(path: Path) -> Optional[str]:
    stem = path.stem

    if stem.lower().startswith("chart - "):
        title = stem[8:]

        if " _ bestdori" in title.lower():
            title = re.split(r"\s+_\s+bestdori", title, flags=re.IGNORECASE)[0]

        # REMOVE difficulty tag like [Expert]
        title = re.sub(r"\s*\[[^\]]+\]\s*", "", title)

        return title.strip() or None

    return None


def _best_effort_json_payload(data: Any, path: Path) -> Dict[str, Any]:
    """
    Build a minimal canonical payload from raw JSON-like Bandori sources.

    This function is intentionally conservative and does not assume a fixed schema.
    """
    title = None
    difficulty = None
    bpm = 0.0
    max_time_beats = 0.0
    note_events = []

    if isinstance(data, dict):
        title = data.get("title") or data.get("name") or data.get("song_name")
        difficulty = data.get("difficulty_label") or data.get("difficulty") or data.get("diff")

        chart = data.get("chart")
        if chart is None and isinstance(data.get("post"), dict):
            chart = data["post"].get("chart")

        if isinstance(chart, list):
            for item in chart:
                if not isinstance(item, dict):
                    continue
                beat = item.get("beat") or item.get("time") or item.get("b")
                lane = item.get("lane") or item.get("track") or item.get("l")
                if isinstance(beat, (int, float)) and isinstance(lane, (int, float)):
                    note_events.append({
                        "time_beats": float(beat),
                        "lane": float(lane),
                        "kind": "tap",
                        "extra": {},
                    })
            if note_events:
                max_time_beats = max(ev["time_beats"] for ev in note_events)

        bpm_val = data.get("bpm")
        if isinstance(bpm_val, (int, float)):
            bpm = float(bpm_val)

    if difficulty is None:
        difficulty = _infer_difficulty_from_name(path) or "unknown"

    if title is None:
        title = _infer_title_from_bestdori_name(path)

    payload: Dict[str, Any] = {
        "game_id": "bandori",
        "chart_id": _infer_chart_id_from_path(path),
        "difficulty": str(difficulty),
        "note_events": note_events,
        "chart_meta": {
            "bpm": float(bpm),
            "max_time_beats": float(max_time_beats),
        },
        "diagnostics": {
            "routing_only": False,
            "source_format": "json",
            "parsing_mode": "best_effort_json",
            "note_event_count": len(note_events),
        },
    }

    if title:
        payload["diagnostics"]["title_inferred"] = title

    return payload


def _norm_lookup_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().casefold()
    text = re.sub(r"\s+", " ", text)
    return text


def _safe_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _safe_number(value: Any) -> Any:
    # Keep Excel cell values as-is unless empty
    if value in (None, ""):
        return None
    return value


@lru_cache(maxsize=4)
def _load_bandori_song_db_index(
    db_path_str: str,
    sheet_name: str = DEFAULT_BANDORI_SHEET_NAME,
) -> Dict[str, Dict[Any, Dict[str, Any]]]:
    """
    Build a cached lookup index from Song Database workbook.

    Returns:
        {
            "by_song_id_diff": {(song_id, diff_code): row_dict, ...},
            "by_name_diff": {(normalized_name, diff_code): row_dict, ...},
        }
    """
    db_path = Path(db_path_str)

    if not db_path.exists():
        return {
            "by_song_id_diff": {},
            "by_name_diff": {},
        }

    wb = load_workbook(db_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {
            "by_song_id_diff": {},
            "by_name_diff": {},
        }

    ws = wb[sheet_name]

    # Read header row
    if ws.max_row < 1:
        return {
            "by_song_id_diff": {},
            "by_name_diff": {},
        }

    header = [cell.value for cell in ws[1]]
    if not header:
        return {
            "by_song_id_diff": {},
            "by_name_diff": {},
        }

    by_song_id_diff: Dict[tuple[Any, Any], Dict[str, Any]] = {}
    by_name_diff: Dict[tuple[Any, Any], Dict[str, Any]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        row_dict: Dict[str, Any] = {}
        for idx, col_name in enumerate(header):
            if col_name is None:
                continue
            row_dict[str(col_name)] = row[idx] if idx < len(row) else None

        song_id = _safe_str(row_dict.get("song_id"))
        name = _safe_str(row_dict.get("name"))
        diff_code = _safe_str(row_dict.get("difficulty_code"))

        if song_id:
            by_song_id_diff[(song_id, diff_code)] = row_dict

        if name:
            by_name_diff[(_norm_lookup_text(name), diff_code)] = row_dict

    return {
        "by_song_id_diff": by_song_id_diff,
        "by_name_diff": by_name_diff,
    }


def _lookup_bandori_song_meta(
    *,
    song_id: Optional[str],
    title: Optional[str],
    difficulty_code: Optional[str],
    db_path: Optional[str | Path] = None,
    sheet_name: str = DEFAULT_BANDORI_SHEET_NAME,
) -> Optional[Dict[str, Any]]:
    """
    Complementary Song DB lookup.

    Preference:
    1) (song_id, difficulty_code)
    2) (normalized title, difficulty_code)
    3) (song_id, None)
    4) (normalized title, None)
    """
    db_path_resolved = Path(db_path) if db_path else Path.cwd() / DEFAULT_SONG_DB_FILENAME
    index = _load_bandori_song_db_index(str(db_path_resolved), sheet_name=sheet_name)

    by_song_id_diff = index["by_song_id_diff"]
    by_name_diff = index["by_name_diff"]

    # Primary exact match: song_id + difficulty
    if song_id:
        hit = by_song_id_diff.get((_safe_str(song_id), _safe_str(difficulty_code)))
        if hit:
            return hit

    # Secondary exact match: normalized title + difficulty
    if title:
        hit = by_name_diff.get((_norm_lookup_text(title), _safe_str(difficulty_code)))
        if hit:
            return hit

    # Fallback: song_id without difficulty
    if song_id:
        hit = by_song_id_diff.get((_safe_str(song_id), None))
        if hit:
            return hit

    # Fallback: title without difficulty
    if title:
        hit = by_name_diff.get((_norm_lookup_text(title), None))
        if hit:
            return hit

    return None


def _prefer_primary(primary_value: Any, fallback_value: Any) -> Any:
    """
    Keep primary chart-derived value if present; otherwise use Song DB fallback.
    """
    if primary_value not in (None, ""):
        return primary_value
    return fallback_value

class BandoriAdapter(BaseAdapter):
    """Bandori adapter for UMI routing + minimal canonicalization."""

    game_id: str = "bandori"
    adapter_id: str = "adapter_bandori"
    adapter_version: str = "1.1.0"

    def accepts_file(self, path: Any) -> bool:
        p = Path(str(path))
        return p.suffix.lower() in {".json", ".html", ".mht"}

    def parse_file(self, path: Path):
        # Infer basic fields
        chart_id = _infer_chart_id_from_path(path)
        difficulty = _infer_difficulty_from_name(path)
        title = _infer_title_from_bestdori_name(path)

        payload = {
            "game_id": "bandori",
            "chart_id": chart_id,
            "title": title,
            "difficulty": difficulty or "UNKNOWN",
            "source_file": str(path.name),
        }

        return payload

    
    def to_canonical_row(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        diagnostics = payload.get("diagnostics", {}) if isinstance(payload.get("diagnostics"), dict) else {}
        chart_meta = payload.get("chart_meta", {}) if isinstance(payload.get("chart_meta"), dict) else {}
        adapter_meta = payload.get("adapter_metadata", {}) if isinstance(payload.get("adapter_metadata"), dict) else {}

        # ----- core fields -----
        title = payload.get("title") or diagnostics.get("title_inferred")

        source_path = (
            payload.get("source_file")
            or adapter_meta.get("source_path")
            or payload.get("source_path")
        )

        difficulty = payload.get("difficulty")

        # ----- derived fields from payload -----
        note_events = payload.get("note_events") if isinstance(payload.get("note_events"), list) else []
        note_total_chart = len(note_events) if note_events else None

        bpm = chart_meta.get("bpm")
        max_beats = chart_meta.get("max_time_beats")

        duration_ms = None
        if isinstance(bpm, (int, float)) and bpm > 0 and isinstance(max_beats, (int, float)):
            duration_ms = int((max_beats / bpm) * 60 * 1000)

        # Lightweight level inference (keep as current safe heuristic)
        level = None
        if isinstance(difficulty, str):
            diff = difficulty.upper()
            if diff == "EASY":
                level = 5
            elif diff == "NORMAL":
                level = 10
            elif diff == "HARD":
                level = 18
            elif diff == "EXPERT":
                level = 25
            elif diff == "SPECIAL":
                level = 30

        # ----- complementary Song DB lookup -----
        db_hit = _lookup_bandori_song_meta(
            song_id=_safe_str(payload.get("chart_id")),
            title=_safe_str(title),
            difficulty_code=_safe_str(difficulty),
            db_path=None,   # uses default: Song Database (full).xlsx in cwd
            sheet_name=DEFAULT_BANDORI_SHEET_NAME,
        )

        # ----- merge strategy -----
        # Keep payload/chart-derived values as the primary source.
        # Only use DB values to fill gaps.
        return {
            "game_id": self.game_id,

            # identity
            "song_id": _prefer_primary(payload.get("chart_id"), db_hit.get("song_id") if db_hit else None),
            "name": _prefer_primary(title, db_hit.get("name") if db_hit else None),

            # difficulty
            "difficulty_label": _prefer_primary(difficulty, db_hit.get("difficulty_label") if db_hit else None),
            "difficulty_code": _prefer_primary(difficulty, db_hit.get("difficulty_code") if db_hit else None),

            # structure
            "tier": _prefer_primary(None, db_hit.get("tier") if db_hit else None),
            "level": _prefer_primary(level, _safe_number(db_hit.get("level")) if db_hit else None),
            "rating_raw": _prefer_primary(None, _safe_number(db_hit.get("rating_raw")) if db_hit else None),

            # notes
            "note_total_db": _prefer_primary(None, _safe_number(db_hit.get("note_total_db")) if db_hit else None),
            "note_total_chart": _prefer_primary(note_total_chart, _safe_number(db_hit.get("note_total_chart")) if db_hit else None),
            "note_delta": _prefer_primary(None, _safe_number(db_hit.get("note_delta")) if db_hit else None),

            # file
            "chart_path": _prefer_primary(source_path, db_hit.get("chart_path") if db_hit else None),

            # timing
            "duration_ms": _prefer_primary(duration_ms, _safe_number(db_hit.get("duration_ms")) if db_hit else None),
            "bpm": _prefer_primary(bpm, _safe_number(db_hit.get("bpm")) if db_hit else None),
        }



    def load(self, path: Any) -> Dict[str, Any]:
        """
        Minimal source loader.

        Current behavior:
        - .json: parse JSON best-effort
        - .html / .mht: keep raw text for heuristics / future parser work
        """
        p = Path(str(path))
        suffix = p.suffix.casefold()

        if suffix == ".json":
            return {
                "path": str(p),
                "source_format": "json",
                "raw": json.loads(p.read_text(encoding="utf-8")),
            }

        if suffix in {".html", ".mht"}:
            text = p.read_text(encoding="utf-8", errors="ignore")
            return {
                "path": str(p),
                "source_format": suffix.lstrip("."),
                "raw": None,
                "text": text,
            }

        raise ValueError(f"Unsupported Bandori source format: {p}")

    def to_canonical_payload(self, path: str) -> Dict[str, Any]:
        raw = self.load(path)
        p = Path(path)
        source_format = str(raw.get("source_format") or p.suffix.lstrip("."))

        if source_format == "json":
            payload = _best_effort_json_payload(raw.get("raw"), p)
        else:
            title = _infer_title_from_bestdori_name(p)
            difficulty = _infer_difficulty_from_name(p) or "unknown"
            text = raw.get("text") if isinstance(raw.get("text"), str) else None

            payload = {
                "game_id": self.game_id,
                "chart_id": _infer_chart_id_from_path(p),
                "difficulty": difficulty,
                "note_events": [],
                "chart_meta": {
                    "bpm": 0.0,
                    "max_time_beats": 0.0,
                },
                "diagnostics": {
                    "routing_only": True,
                    "source_format": source_format,
                    "parsing_implemented": False,
                    "bandori_export_hint": _looks_like_bandori_export(p, text),
                },
            }

            if title:
                payload["diagnostics"]["title_inferred"] = title

        meta_block = build_adapter_metadata(
            adapter_id=self.adapter_id,
            adapter_version=self.adapter_version,
            source_format=source_format,
            source_path=str(p),
            notes=(
                "Bandori fallback-capable adapter. "
                "JSON is best-effort parsed; HTML/MHT currently route with minimal payload."
            ),
        )
        attach_adapter_metadata(payload, meta_block)

        attach_internal_metadata(
            payload,
            adapter_id=self.adapter_id,
            adapter_version=self.adapter_version,
            sections_source="bandori-fallback-adapter",
        )

        return payload

    def capabilities(self) -> Dict[str, Any]:
        return {
            "note_model": "unknown",
            "supports_sections": False,
            "supports_variable_bpm": False,
            "supports_ground_truth_chart": False,
            "emits_canonical_payload": True,
            "routing_only_html_mht": True,
        }


__all__ = ["BandoriAdapter"]