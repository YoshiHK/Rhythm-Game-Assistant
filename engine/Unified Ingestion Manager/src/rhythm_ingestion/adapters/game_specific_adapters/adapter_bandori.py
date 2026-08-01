#!/usr/bin/env python3
from __future__ import annotations

"""
adapter_bandori.py

Bandori adapter (fallback-capable, v2-normalized) for UMI.

Purpose:
- keep the adapter import-safe when legacy bandori_model / bandori_chart modules are absent
- support routing-level acceptance for .json plus baseline fallback .html/.mht
- emit a minimal but valid canonical payload so UMI can produce rows
- preserve a future path for richer Bestdori/HTML parsing without blocking routing today

IMPORTANT:
- .json remains the primary machine-readable source format.
- .html / .mht support is heuristic and filename/text based.
- This adapter intentionally avoids gameplay semantics, tips generation, and heavy inference.
"""

import json
import re
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from ..base_adapter_v2 import BaseAdapterV2
from ..common_adapter_utils import (
    attach_if_missing,
    build_internal_metadata,
    with_baseline_fallback_extensions,
)


DEFAULT_SONG_DB_FILENAME = "Song Database (full).xlsx"
DEFAULT_BANDORI_SHEET_NAME = "BanG Dream"

_BESTDORI_HINT_TOKENS = {
    "bestdori",
    "bang dream",
    "gbp resource site",
    "bandori",
    "garupa",
}

_DIFFICULTY_TOKENS = [
    "special",
    "expert",
    "hard",
    "normal",
    "easy",
]

_DIFFICULTY_UPPER = {
    "easy": "EASY",
    "normal": "NORMAL",
    "hard": "HARD",
    "expert": "EXPERT",
    "special": "SPECIAL",
    "unknown": "UNKNOWN",
}


# ---------------------------------------------------------------------
# Path / filename heuristics
# ---------------------------------------------------------------------
def _looks_like_bandori_export(path: Path, text: Optional[str] = None) -> bool:
    name = path.name.casefold()
    if any(tok in name for tok in _BESTDORI_HINT_TOKENS):
        return True
    if text:
        lowered = text.casefold()
        return any(tok in lowered for tok in _BESTDORI_HINT_TOKENS)
    return False


def _looks_like_bandori_path(path: Path) -> bool:
    s = str(path).replace("\\", "/").casefold()
    name = path.name.casefold()

    path_hints = (
        "bandori",
        "bestdori",
        "bang dream",
        "garupa",
    )

    return any(h in s for h in path_hints) or any(h in name for h in path_hints)


def _infer_difficulty_from_name(path: Path) -> Optional[str]:
    # 1) filename tokens
    stem = path.stem.casefold()
    for token in _DIFFICULTY_TOKENS:
        if f"[{token}]" in stem or token in stem:
            return token.upper()

    # 2) parent folders
    for part in reversed(path.parts):
        lowered = str(part).strip().casefold()
        if lowered in _DIFFICULTY_UPPER:
            return _DIFFICULTY_UPPER[lowered]

    return None


def _infer_chart_id_from_path(path: Path) -> str:
    return path.stem.strip() or path.name


def _infer_title_from_bestdori_name(path: Path) -> Optional[str]:
    stem = path.stem

    if stem.lower().startswith("chart - "):
        title = stem[8:]

        if " _ bestdori" in title.lower():
            title = re.split(r"\s+_\s+bestdori", title, flags=re.IGNORECASE)[0]

        # remove difficulty tag like [Expert]
        title = re.sub(r"\s*\[[^\]]+\]\s*", "", title)

        return title.strip() or None

    return None


def _infer_title_generic(path: Path) -> Optional[str]:
    title = _infer_title_from_bestdori_name(path)
    if title:
        return title

    stem = path.stem.strip()
    if not stem:
        return None

    # remove obvious trailing Bestdori suffix if present
    stem = re.split(r"\s+_\s+Bestdori", stem, flags=re.IGNORECASE)[0]
    stem = re.sub(r"\s*\[[^\]]+\]\s*", "", stem).strip()
    return stem or None


# ---------------------------------------------------------------------
# Safe conversions
# ---------------------------------------------------------------------
def _safe_int(x: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        if x is None:
            return default
        return int(float(x))
    except Exception:
        return default


def _safe_float(x: Any, default: Optional[float] = None) -> Optional[float]:
    try:
        if x is None:
            return default
        return float(x)
    except Exception:
        return default


def _norm_lookup_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().casefold()
    text = re.sub(r"\s+", " ", text)
    return text


def _safe_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _safe_number(value: Any) -> Any:
    if value in (None, ""):
        return None
    return value


def _normalize_difficulty_label(value: Any) -> Optional[str]:
    s = _safe_str(value)
    if not s:
        return None
    lowered = s.casefold()
    return _DIFFICULTY_UPPER.get(lowered, s.upper())


# ---------------------------------------------------------------------
# Song DB lookup
# ---------------------------------------------------------------------
@lru_cache(maxsize=4)
def _load_bandori_song_db_index(
    db_path_str: str,
    sheet_name: str = DEFAULT_BANDORI_SHEET_NAME,
) -> Dict[str, Dict[Any, Dict[str, Any]]]:
    """
    Build a cached lookup index from Song Database workbook.

    Returns:
        {
            "by_song_id_diff": {(song_id, diff_code): row_dict, ...},
            "by_name_diff": {(normalized_name, diff_code): row_dict, ...},
        }
    """
    db_path = Path(db_path_str)

    if not db_path.exists():
        return {
            "by_song_id_diff": {},
            "by_name_diff": {},
        }

    wb = load_workbook(db_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {
            "by_song_id_diff": {},
            "by_name_diff": {},
        }

    ws = wb[sheet_name]

    if ws.max_row < 1:
        return {
            "by_song_id_diff": {},
            "by_name_diff": {},
        }

    header = [cell.value for cell in ws[1]]
    if not header:
        return {
            "by_song_id_diff": {},
            "by_name_diff": {},
        }

    by_song_id_diff: Dict[Tuple[Any, Any], Dict[str, Any]] = {}
    by_name_diff: Dict[Tuple[Any, Any], Dict[str, Any]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        row_dict: Dict[str, Any] = {}
        for idx, col_name in enumerate(header):
            if col_name is None:
                continue
            row_dict[str(col_name)] = row[idx] if idx < len(row) else None

        song_id = _safe_str(row_dict.get("song_id"))
        name = _safe_str(row_dict.get("name"))
        diff_code = _safe_str(row_dict.get("difficulty_code"))

        if song_id:
            by_song_id_diff[(song_id, diff_code)] = row_dict

        if name:
            by_name_diff[(_norm_lookup_text(name), diff_code)] = row_dict

    return {
        "by_song_id_diff": by_song_id_diff,
        "by_name_diff": by_name_diff,
    }


def _lookup_bandori_song_meta(
    *,
    song_id: Optional[str],
    title: Optional[str],
    difficulty_code: Optional[str],
    db_path: Optional[str | Path] = None,
    sheet_name: str = DEFAULT_BANDORI_SHEET_NAME,
) -> Optional[Dict[str, Any]]:
    """
    Complementary Song DB lookup.

    Preference:
    1) (song_id, difficulty_code)
    2) (normalized title, difficulty_code)
    3) (song_id, None)
    4) (normalized title, None)
    """
    db_path_resolved = Path(db_path) if db_path else Path.cwd() / DEFAULT_SONG_DB_FILENAME
    index = _load_bandori_song_db_index(str(db_path_resolved), sheet_name=sheet_name)

    by_song_id_diff = index["by_song_id_diff"]
    by_name_diff = index["by_name_diff"]

    if song_id:
        hit = by_song_id_diff.get((_safe_str(song_id), _safe_str(difficulty_code)))
        if hit:
            return hit

    if title:
        hit = by_name_diff.get((_norm_lookup_text(title), _safe_str(difficulty_code)))
        if hit:
            return hit

    if song_id:
        hit = by_song_id_diff.get((_safe_str(song_id), None))
        if hit:
            return hit

    if title:
        hit = by_name_diff.get((_norm_lookup_text(title), None))
        if hit:
            return hit

    return None


def _prefer_primary(primary_value: Any, fallback_value: Any) -> Any:
    if primary_value not in (None, ""):
        return primary_value
    return fallback_value


# ---------------------------------------------------------------------
# Payload builders
# ---------------------------------------------------------------------
def _best_effort_json_payload(data: Any, path: Path) -> Dict[str, Any]:
    """
    Build a minimal canonical payload from raw JSON-like Bandori sources.

    This function is intentionally conservative and does not assume a fixed schema.
    """
    title = None
    difficulty = None
    bpm = 0.0
    max_time_beats = 0.0
    note_events: List[Dict[str, Any]] = []

    if isinstance(data, dict):
        title = data.get("title") or data.get("name") or data.get("song_name")
        difficulty = data.get("difficulty_label") or data.get("difficulty") or data.get("diff")

        chart = data.get("chart")
        if chart is None and isinstance(data.get("post"), dict):
            chart = data["post"].get("chart")

        if isinstance(chart, list):
            for item in chart:
                if not isinstance(item, dict):
                    continue
                beat = item.get("beat") or item.get("time") or item.get("b")
                lane = item.get("lane") or item.get("track") or item.get("l")
                if isinstance(beat, (int, float)) and isinstance(lane, (int, float)):
                    note_events.append({
                        "time_beats": float(beat),
                        "lane": float(lane),
                        "kind": "tap",
                        "extra": {},
                    })
            if note_events:
                max_time_beats = max(ev["time_beats"] for ev in note_events)

        bpm_val = data.get("bpm")
        if isinstance(bpm_val, (int, float)):
            bpm = float(bpm_val)

    if difficulty is None:
        difficulty = _infer_difficulty_from_name(path) or "UNKNOWN"
    else:
        difficulty = _normalize_difficulty_label(difficulty) or "UNKNOWN"

    if title is None:
        title = _infer_title_generic(path)

    payload: Dict[str, Any] = {
        "game_id": "bandori",
        "chart_id": _infer_chart_id_from_path(path),
        "difficulty": str(difficulty),
        "title": title,
        "note_events": note_events,
        "chart_meta": {
            "bpm": float(bpm),
            "max_time_beats": float(max_time_beats),
        },
        "diagnostics": {
            "routing_only": False,
            "source_format": "json",
            "parsing_mode": "best_effort_json",
            "note_event_count": len(note_events),
        },
        "adapter_metadata": {
            "adapter_id": "adapter_bandori",
            "adapter_version": "2.1.0",
            "source_format": "json",
            "source_path": str(path),
            "notes": "Bandori adapter parsing JSON with best-effort structural mapping.",
        },
    }

    if title:
        payload["diagnostics"]["title_inferred"] = title

    return payload


def _best_effort_html_payload(text: str, path: Path) -> Dict[str, Any]:
    """
    Build a minimal canonical payload from Bestdori-like HTML / MHT sources.

    This is intentionally lightweight:
    - infer title from filename
    - infer difficulty from filename/folders
    - do not attempt gameplay-semantic parsing
    - note_events are empty unless a future richer parser is added
    """
    title = _infer_title_generic(path)
    difficulty = _infer_difficulty_from_name(path) or "UNKNOWN"

    payload: Dict[str, Any] = {
        "game_id": "bandori",
        "chart_id": _infer_chart_id_from_path(path),
        "difficulty": difficulty,
        "title": title,
        "note_events": [],
        "chart_meta": {
            "bpm": 0.0,
            "max_time_beats": 0.0,
        },
        "diagnostics": {
            "routing_only": False,
            "source_format": path.suffix.lower().lstrip("."),
            "parsing_mode": "best_effort_html",
            "bestdori_detected": _looks_like_bandori_export(path, text),
            "note_event_count": 0,
        },
        "adapter_metadata": {
            "adapter_id": "adapter_bandori",
            "adapter_version": "2.1.0",
            "source_format": path.suffix.lower().lstrip("."),
            "source_path": str(path),
            "notes": "Bandori adapter using filename/text heuristics for HTML/MHT fallback.",
        },
    }

    if title:
        payload["diagnostics"]["title_inferred"] = title

    return payload


# ---------------------------------------------------------------------
# Adapter class
# ---------------------------------------------------------------------
class BandoriAdapter(BaseAdapterV2):
    game_id = "bandori"
    adapter_id = "adapter_bandori"
    adapter_version = "2.1.0"

    def accepts_file(self, path) -> bool:
        p = Path(path)
        allowed = with_baseline_fallback_extensions([".json"])
        if p.suffix.lower() not in allowed:
            return False
        return _looks_like_bandori_path(p)

    def load(self, path):
        p = Path(path)
        suffix = p.suffix.lower()

        if suffix == ".json":
            try:
                return json.loads(p.read_text(encoding="utf-8", errors="ignore"))
            except Exception:
                # keep the raw text as last-resort fallback if malformed JSON
                return p.read_text(encoding="utf-8", errors="ignore")

        # html / mht / others supported by baseline fallback
        return p.read_text(encoding="utf-8", errors="ignore")

    def _enrich_payload_from_song_db(
        self,
        payload: Dict[str, Any],
        *,
        path: Path,
    ) -> Dict[str, Any]:
        """
        Optional Song DB enrichment:
        - best-effort only
        - never blocks routing
        - chart-derived values remain primary
        """
        song_id = _safe_str(payload.get("chart_id"))
        title = _safe_str(payload.get("title"))
        difficulty_code = _safe_str(payload.get("difficulty"))

        hit = _lookup_bandori_song_meta(
            song_id=song_id,
            title=title,
            difficulty_code=difficulty_code,
        )

        if not hit:
            return payload

        # Preserve chart-derived values if already present; use Song DB as fallback.
        payload["title"] = _prefer_primary(payload.get("title"), _safe_str(hit.get("name")))
        payload["difficulty"] = _prefer_primary(
            payload.get("difficulty"),
            _normalize_difficulty_label(hit.get("difficulty_label") or hit.get("difficulty_code")),
        )

        payload["diagnostics"]["song_db_hit"] = True
        payload["diagnostics"]["song_db_match_key"] = {
            "song_id": song_id,
            "title": title,
            "difficulty_code": difficulty_code,
        }

        attach_if_missing(
            payload,
            "song_db_metadata",
            {
                "song_id": _safe_str(hit.get("song_id")),
                "name": _safe_str(hit.get("name")),
                "title_en": _safe_str(hit.get("title_en")),
                "difficulty_code": _safe_str(hit.get("difficulty_code")),
                "difficulty_label": _safe_str(hit.get("difficulty_label")),
                "level": _safe_number(hit.get("level")),
                "bpm": _safe_number(hit.get("bpm")),
            },
        )

        # only use Song DB BPM if chart_meta.bpm is still zero / empty
        chart_meta = payload.get("chart_meta")
        if isinstance(chart_meta, dict):
            current_bpm = _safe_float(chart_meta.get("bpm"), 0.0) or 0.0
            if current_bpm <= 0:
                db_bpm = _safe_float(hit.get("bpm"), None)
                if db_bpm is not None:
                    chart_meta["bpm"] = float(db_bpm)

        return payload

    def to_canonical_payload(self, path: str) -> Dict[str, Any]:
        p = Path(path)
        raw = self.load(p)

        if isinstance(raw, dict):
            payload = _best_effort_json_payload(raw, p)
        else:
            text = str(raw) if raw is not None else ""
            payload = _best_effort_html_payload(text, p)

        payload = self._enrich_payload_from_song_db(payload, path=p)

        # Standard v2 finalization:
        # guarantee minimal payload contract + internal metadata
        payload = self.finalize_payload_v2(
            payload,
            source_path=str(p),
            default_chart_id=payload.get("chart_id") or _infer_chart_id_from_path(p),
            default_difficulty=payload.get("difficulty") or _infer_difficulty_from_name(p) or "UNKNOWN",
        )

        # Attach enriched internal metadata in an additive way
        internal = build_internal_metadata(
            adapter_id=self.adapter_id,
            adapter_version=self.adapter_version,
            sections_source=None,
            notes="Bandori adapter v2 finalized payload",
            extra={
                "source_file_suffix": p.suffix.lower(),
                "fallback_mode": p.suffix.lower() in {".html", ".mht"},
            },
        )
        attach_if_missing(payload, "internal_metadata", internal)

        return payload

    def to_canonical_row(self, raw) -> Dict[str, Any]:
        """
        Accept either:
        - canonical payload dict
        - path-like object / string
        """
        if isinstance(raw, dict):
            payload = raw
        else:
            payload = self.to_canonical_payload(str(raw))

        chart_meta = payload.get("chart_meta") if isinstance(payload, dict) else {}
        if not isinstance(chart_meta, dict):
            chart_meta = {}

        note_events = payload.get("note_events") if isinstance(payload, dict) else []
        if not isinstance(note_events, list):
            note_events = []

        title = _safe_str(payload.get("title")) or _safe_str(payload.get("chart_id")) or "UNKNOWN"

        difficulty = _normalize_difficulty_label(payload.get("difficulty")) or "UNKNOWN"
        song_id = _safe_str(payload.get("chart_id")) or title

        bpm = _safe_float(chart_meta.get("bpm"), None)
        max_time_beats = _safe_float(chart_meta.get("max_time_beats"), 0.0) or 0.0

        # Best-effort duration estimate from beats + bpm
        duration_ms: int = 0
        if bpm is not None and bpm > 0 and max_time_beats >= 0:
            duration_ms = int((max_time_beats / bpm) * 60_000.0)

        row: Dict[str, Any] = {
            "game": self.game_id,
            "game_id": self.game_id,  # additive alias for compatibility
            "song_id": song_id,
            "name": title,
            "title": title,           # additive alias for compatibility
            "tier": difficulty,
            "difficulty_label": difficulty,
            "note_total_chart": len(note_events),
            "note_total_db": None,
            "note_delta": None,
            "duration_ms": duration_ms,
            "bpm": bpm,
        }

        return row

    def capabilities(self) -> Dict[str, Any]:
        return {
            "source_formats": [".json", ".html", ".mht"],
            "primary_format": ".json",
            "supports_best_effort_html": True,
            "supports_song_db_enrichment": True,
            "note_model": "lane_based",
            "sections": False,
        }


__all__ = ["BandoriAdapter"]