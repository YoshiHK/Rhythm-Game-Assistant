# -*- coding: utf-8 -*-
"""
(2-4.1) chart_visual_detector_merged.py

Fully merged chart visual detector for the Project Sekai tips pipeline.

Responsibilities (Steps 2–4.1 in the workflow):
  - Parse Proseka Trainer HTML/SVG exports.
  - Extract a normalized representation of note events (NoteEvent).
  - Look up song metadata (BPM, duration, etc.) from primary Excel database
    (proseka song database.xlsx), with HTML wiki table
    (収録楽曲 _ プロジェクトセカイ攻略Wiki.html) as fallback
    (using fuzzy title matching).
  - Aggregate SectionMetrics (with rest_ratio, hold_coverage,
    notes_during_hold_ratio, overlap_ratio).
  - Run canonical severity engine (automatic_inference_framework).
  - Detect pattern-signal tags according to Tips Generation Guides.xlsx
    (sheet 2), supplemented by canonical severities from proseka_severity_rules.
"""

from __future__ import annotations
from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple

import os
import re
from xml.etree import ElementTree as ET

import openpyxl

from proseka_severity_rules import (
    SectionMetrics,
    automatic_inference_framework,
    severity_ge,
)

# Filenames for metadata sources
SONG_DB_XLSX = "proseka song database.xlsx"
WIKI_HTML = "収録楽曲 _ プロジェクトセカイ攻略Wiki.html"

# In-memory cache for Excel metadata
_SONG_META_CACHE: Dict[str, Dict[str, Any]] | None = None


# ----------------------------------------------------------------------
# Core datatypes
# ----------------------------------------------------------------------

@dataclass
class NoteEvent:
    """Internal normalized representation of a single note/gesture event."""
    time_beats: float
    lane: int
    kind: str
    extra: Dict[str, Any]


# ----------------------------------------------------------------------
# Public entrypoint
# ----------------------------------------------------------------------

def analyze_chart_html(html_path: str) -> Dict[str, Any]:
    """Analyze a Proseka Trainer HTML export and return structured analysis."""
    svg_root = load_svg_from_html(html_path)
    notes = parse_svg_to_note_events(svg_root)

    # Infer song title + difficulty from filename
    title, diff_name, diff_level = infer_title_and_difficulty_from_filename(html_path)

    # Lookup BPM + duration via Excel first, then wiki fallback (with fuzzy title matching)
    bpm, duration_sec = lookup_song_metadata(title)
    if bpm is None:
        bpm = 120.0  # safe fallback

    sections = build_section_metrics(notes, bpm)
    severity_fw = automatic_inference_framework(sections)
    detected_tags = detect_pattern_tags(notes, bpm, sections, severity_fw)

    meta = {
        "song_title": title,
        "difficulty_name": diff_name,
        "difficulty_level": diff_level,
        "engine_chart_id": None,
        "duration_sec": duration_sec,
        "bpm": bpm,
        "source": "proseka-trainer-html",
    }

    diagnostics = {
        "raw_note_events_count": len(notes),
        "section_count": len(sections),
        "canonical_severities": severity_fw.get("aggregated", {}),
    }

    return {
        "meta": meta,
        "sections": sections,
        "detected_tags": detected_tags,
        "diagnostics": diagnostics,
    }


# ----------------------------------------------------------------------
# Song metadata lookup (Excel primary, Wiki HTML fallback with fuzzy match)
# ----------------------------------------------------------------------

def infer_title_and_difficulty_from_filename(path: str) -> Tuple[str, Optional[str], Optional[int]]:
    """Infer song title and difficulty from a Trainer export filename.

    Example::

        '群青讃歌 (EXPERT 24) の譜面 - プロセカ練習場.html'
        -> title='群青讃歌', diff_name='EXPERT', diff_level=24
    """
    base = os.path.basename(path)
    if base.lower().endswith(".html"):
        base = base[:-5]
    base = base.split(" の譜面")[0]

    if " (" in base and ")" in base:
        title_part, rest = base.split(" (", 1)
        rest = rest.split(")", 1)[0]
        parts = rest.strip().split()
        if len(parts) >= 2:
            diff_name = parts[0]
            try:
                diff_level = int(parts[1])
            except ValueError:
                diff_level = None
        else:
            diff_name = rest.strip()
            diff_level = None
        title = title_part.strip()
    else:
        title = base.strip()
        diff_name = None
        diff_level = None

    return title, diff_name, diff_level


def normalize_title_for_match(s: str) -> str:
    """Normalize song titles for fuzzy matching.

    - lowercased
    - remove spaces (including full-width)
    - remove punctuation commonly appearing in PJSekai titles
    """
    if not s:
        return ""
    s = s.lower()
    remove_chars = " 　!！?？・・/／\"'()[]＜＞<>『』【】."
    for ch in remove_chars:
        s = s.replace(ch, "")
    return s


def simple_similarity(a: str, b: str) -> float:
    """Compute a simple character-overlap similarity between two strings.

    score = 2 * |common_chars| / (len(a) + len(b))
    """
    if not a or not b:
        return 0.0
    set_a = set(a)
    set_b = set(b)
    common = len(set_a & set_b)
    return 2.0 * common / (len(a) + len(b))


def load_song_meta_excel() -> Dict[str, Dict[str, Any]]:
    """Load song metadata from the Excel database into a dict keyed by title."""
    global _SONG_META_CACHE
    if _SONG_META_CACHE is not None:
        return _SONG_META_CACHE

    meta: Dict[str, Dict[str, Any]] = {}
    if not os.path.isfile(SONG_DB_XLSX):
        _SONG_META_CACHE = meta
        return meta

    wb = openpyxl.load_workbook(SONG_DB_XLSX, data_only=True)
    sh = wb[wb.sheetnames[0]]

    header = next(sh.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {name: idx for idx, name in enumerate(header)}
    idx_title = col_map.get("楽曲名")
    idx_time = col_map.get("時間")
    idx_bpm = col_map.get("BPM")

    if idx_title is None or idx_bpm is None:
        _SONG_META_CACHE = meta
        return meta

    for row in sh.iter_rows(min_row=2, values_only=True):
        title = row[idx_title]
        if not title:
            continue
        bpm = row[idx_bpm]
        t = row[idx_time] if idx_time is not None else None
        duration_sec = None
        if hasattr(t, "hour") and hasattr(t, "minute") and hasattr(t, "second"):
            duration_sec = int(t.hour) * 3600 + int(t.minute) * 60 + int(t.second)
        meta[str(title)] = {
            "bpm": float(bpm) if isinstance(bpm, (int, float)) else None,
            "duration_sec": duration_sec,
            "norm_title": normalize_title_for_match(str(title)),
        }

    _SONG_META_CACHE = meta
    return meta


def lookup_song_metadata(title: str) -> Tuple[Optional[float], Optional[float]]:
    """Lookup (bpm, duration_sec) using Excel first, then wiki HTML fallback.

    Uses fuzzy matching on song titles when exact match is not found.
    """
    db = load_song_meta_excel()
    if title in db:
        entry = db[title]
        return entry.get("bpm"), entry.get("duration_sec")

    norm_target = normalize_title_for_match(title)
    best_key = None
    best_score = 0.0
    for key, entry in db.items():
        norm_key = entry.get("norm_title") or normalize_title_for_match(key)
        score = simple_similarity(norm_target, norm_key)
        if score > best_score:
            best_score = score
            best_key = key
    if best_key is not None and best_score >= 0.7:
        entry = db[best_key]
        return entry.get("bpm"), entry.get("duration_sec")

    bpm_html, duration_sec_html = lookup_song_metadata_from_wiki(title)
    return bpm_html, duration_sec_html


def lookup_song_metadata_from_wiki(title: str) -> Tuple[Optional[float], Optional[float]]:
    """Fallback: parse BPM and duration from official wiki HTML table.

    Uses fuzzy matching on the 楽曲名 column of the main songs table.
    """
    if not os.path.isfile(WIKI_HTML):
        return None, None

    with open(WIKI_HTML, "r", encoding="utf-8") as f:
        html = f.read()

    m = re.search(r'<table[^>]+id="sortable_table1"[\s\S]*?</table>', html)
    if not m:
        return None, None
    table_html = m.group(0)

    norm_target = normalize_title_for_match(title)
    best_bpm = None
    best_duration = None
    best_score = 0.0

    row_pattern = re.compile(r"<tr[\s\S]*?</tr>")
    for row_html in row_pattern.findall(table_html):
        cells = re.findall(r"<td[^>]*>([\s\S]*?)</td>", row_html)
        if len(cells) < 16:
            continue
        title_cell = _strip_tags(cells[3]).strip()  # 楽曲名 column
        norm_title_cell = normalize_title_for_match(title_cell)
        score = simple_similarity(norm_target, norm_title_cell)
        if score <= best_score:
            continue

        time_cell = _strip_tags(cells[14]).strip()
        bpm_cell = _strip_tags(cells[15]).strip()

        try:
            bpm_val = float(bpm_cell)
        except ValueError:
            bpm_val = None

        duration_sec = None
        mtime = re.match(r"(\d+):(\d+)", time_cell)
        if mtime:
            mm = int(mtime.group(1))
            ss = int(mtime.group(2))
            duration_sec = mm * 60 + ss

        best_score = score
        best_bpm = bpm_val
        best_duration = duration_sec

    if best_score >= 0.7:
        return best_bpm, best_duration
    return None, None


def _strip_tags(s: str) -> str:
    return re.sub(r"<[^>]+>", "", s)

# ----------------------------------------------------------------------
# HTML/SVG loading and lane/timing helpers
# ----------------------------------------------------------------------

def load_svg_from_html(html_path: str) -> ET.Element:
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()

    start_idx = html.find("<svg")
    if start_idx == -1:
        raise ValueError("No <svg> tag found in HTML.")

    end_idx = html.find("</svg>", start_idx)
    if end_idx == -1:
        raise ValueError("No closing </svg> tag found in HTML.")

    svg_str = html[start_idx : end_idx + len("</svg>")]

    try:
        svg_root = ET.fromstring(svg_str)
    except ET.ParseError as e:
        raise ValueError(f"Failed to parse SVG from HTML: {e}")

    return svg_root


def compute_lane_centers(svg_root: ET.Element) -> List[float]:
    lane_lines: List[float] = []
    for line in svg_root.findall(".//{*}line"):
        x1 = line.get("x1")
        x2 = line.get("x2")
        if x1 is None or x2 is None:
            continue
        try:
            fx1 = float(x1)
            fx2 = float(x2)
        except ValueError:
            continue
        if abs(fx1 - fx2) < 1e-3 and 0.0 <= fx1 <= 100.0:
            lane_lines.append(fx1)
    lane_lines = sorted(_dedupe_floats(lane_lines, tol=0.1))
    if len(lane_lines) < 2:
        raise ValueError("Not enough vertical lane lines found to compute centers.")
    lane_centers: List[float] = []
    for i in range(len(lane_lines) - 1):
        center = (lane_lines[i] + lane_lines[i + 1]) / 2.0
        lane_centers.append(center)
    return lane_centers


def extract_measure_markers(svg_root: ET.Element) -> List[Tuple[int, float]]:
    markers: List[Tuple[int, float]] = []
    for text_el in svg_root.findall(".//{*}text"):
        txt = (text_el.text or "").strip()
        if not txt:
            continue
        if not re.fullmatch(r"-?\d+", txt):
            continue
        try:
            measure_num = int(txt)
        except ValueError:
            continue
        y_attr = text_el.get("y")
        if y_attr is None:
            continue
        try:
            fy = float(y_attr)
        except ValueError:
            continue
        markers.append((measure_num, fy))
    markers.sort(key=lambda t: t[1])
    if not markers:
        raise ValueError("No measure markers found in SVG <text> elements.")
    return markers


def compute_px_per_beat(measure_markers: List[Tuple[int, float]]) -> float:
    if len(measure_markers) < 2:
        return 40.0
    deltas = []
    for (m1, y1), (m2, y2) in zip(measure_markers, measure_markers[1:]):
        if m2 != m1 + 1:
            continue
        deltas.append(abs(y2 - y1))
    if not deltas:
        return 40.0
    avg_px_per_measure = sum(deltas) / len(deltas)
    return avg_px_per_measure / 4.0


def beats_from_y(
    y: float,
    measure_markers: List[Tuple[int, float]],
    px_per_beat: float,
) -> float:
    first_measure, first_y = measure_markers[0]
    beats_at_first_marker = first_measure * 4.0
    dy = first_y - y
    delta_beats = dy / px_per_beat
    return beats_at_first_marker + delta_beats


def lane_from_x(x: float, lane_centers: List[float]) -> int:
    best_idx = 0
    best_dist = float("inf")
    for i, cx in enumerate(lane_centers):
        d = abs(x - cx)
        if d < best_dist:
            best_dist = d
            best_idx = i
    return best_idx


def _dedupe_floats(values: List[float], tol: float = 0.1) -> List[float]:
    if not values:
        return []
    values = sorted(values)
    deduped = [values[0]]
    for v in values[1:]:
        if abs(v - deduped[-1] > tol):
            deduped.append(v)
    return deduped


def _extract_first_xy_from_path_d(d: str) -> Tuple[Optional[float], Optional[float]]:
    tokens = d.replace(",", " ").split()
    if not tokens:
        return None, None
    if tokens[0].upper() == "M" and len(tokens) >= 3:
        try:
            x = float(tokens[1])
            y = float(tokens[2])
            return x, y
        except ValueError:
            return None, None
    nums: List[float] = []
    for t in tokens:
        try:
            nums.append(float(t))
            if len(nums) == 2:
                break
        except ValueError:
            continue
    if len(nums) == 2:
        return nums[0], nums[1]
    return None, None


def _extract_first_xy_from_points(points: str) -> Tuple[Optional[float], Optional[float]]:
    pts = points.strip().split()
    if not pts:
        return None, None
    first = pts[0]
    if "," in first:
        xs, ys = first.split(",", 1)
    elif " " in first:
        xs, ys = first.split(" ", 1)
    else:
        return None, None
    try:
        return float(xs), float(ys)
    except ValueError:
        return None, None


# ----------------------------------------------------------------------
# Color -> note-kind mapping & SVG -> NoteEvent
# ----------------------------------------------------------------------

COLOR_MAP: Dict[str, str] = {
    "#ab94ff": "tap",
    "#4dc775": "hold_body_or_start",
    "#bbbb00": "critical_tap",
    "#ec62b0": "flick",
    "#ed7f9b": "flick_arrow",
}

TRACE_COLOR_MAP: Dict[str, str] = {
    "#9fd9c0": "hold_path",
    "#e3df65": "critical_hold_path",
}


def _normalize_color(c: Optional[str]) -> Optional[str]:
    if c is None:
        return None
    c = c.strip().lower()
    if len(c) == 9 and c.startswith("#"):
        c = c[:7]
    return c


def map_color_to_kind(stroke: Optional[str], fill: Optional[str]) -> Optional[str]:
    s = _normalize_color(stroke)
    f = _normalize_color(fill)
    if s in COLOR_MAP:
        return COLOR_MAP[s]
    if f in COLOR_MAP:
        return COLOR_MAP[f]
    return None


def map_path_color_to_trace_kind(stroke: Optional[str]) -> Optional[str]:
    s = _normalize_color(stroke)
    if s in TRACE_COLOR_MAP:
        return TRACE_COLOR_MAP[s]
    return None


def compute_width_lanes(rect_width: float, lane_centers: List[float]) -> int:
    if len(lane_centers) < 2:
        return 1
    lane_spacing = abs(lane_centers[1] - lane_centers[0])
    if lane_spacing <= 0:
        return 1
    return max(1, int(round(rect_width / lane_spacing)))


def parse_svg_to_note_events(svg_root: ET.Element) -> List[NoteEvent]:
    lane_centers = compute_lane_centers(svg_root)
    measure_markers = extract_measure_markers(svg_root)
    px_per_beat = compute_px_per_beat(measure_markers)

    events: List[NoteEvent] = []

    for rect in svg_root.findall(".//{*}rect"):
        stroke = rect.get("stroke")
        fill = rect.get("fill")
        x_attr = rect.get("x")
        y_attr = rect.get("y")
        w_attr = rect.get("width")
        h_attr = rect.get("height")
        if x_attr is None or y_attr is None or w_attr is None or h_attr is None:
            continue
        try:
            x = float(x_attr)
            y = float(y_attr)
            w = float(w_attr)
            h = float(h_attr)
        except ValueError:
            continue
        kind = map_color_to_kind(stroke, fill)
        if kind is None:
            continue
        lane_idx = lane_from_x(x, lane_centers)
        time_beats = beats_from_y(y, measure_markers, px_per_beat)
        extra = {
            "raw_stroke": stroke,
            "raw_fill": fill,
            "width_lanes": compute_width_lanes(w, lane_centers),
            "rect_height": h,
            "shape": "rect",
        }
        events.append(NoteEvent(time_beats=time_beats, lane=lane_idx, kind=kind, extra=extra))

    for circle in svg_root.findall(".//{*}circle"):
        stroke = circle.get("stroke")
        fill = circle.get("fill")
        cx_attr = circle.get("cx")
        cy_attr = circle.get("cy")
        if cx_attr is None or cy_attr is None:
            continue
        try:
            x = float(cx_attr)
            y = float(cy_attr)
        except ValueError:
            continue
        kind = map_color_to_kind(stroke, fill)
        if kind is None:
            continue
        lane_idx = lane_from_x(x, lane_centers)
        time_beats = beats_from_y(y, measure_markers, px_per_beat)
        extra = {
            "raw_stroke": stroke,
            "raw_fill": fill,
            "shape": "circle",
        }
        events.append(NoteEvent(time_beats=time_beats, lane=lane_idx, kind=kind, extra=extra))

    for path in svg_root.findall(".//{*}path"):
        stroke = path.get("stroke")
        kind = map_path_color_to_trace_kind(stroke)
        if kind is None:
            continue
        d = path.get("d") or ""
        x, y = _extract_first_xy_from_path_d(d)
        if x is None or y is None:
            continue
        lane_idx = lane_from_x(x, lane_centers)
        time_beats = beats_from_y(y, measure_markers, px_per_beat)
        extra = {
            "raw_stroke": stroke,
            "shape": "path",
            "d": d,
        }
        events.append(NoteEvent(time_beats=time_beats, lane=lane_idx, kind=kind, extra=extra))

    for poly in svg_root.findall(".//{*}polyline"):
        stroke = poly.get("stroke")
        kind = map_path_color_to_trace_kind(stroke)
        if kind is None:
            continue
        points = poly.get("points") or ""
        x, y = _extract_first_xy_from_points(points)
        if x is None or y is None:
            continue
        lane_idx = lane_from_x(x, lane_centers)
        time_beats = beats_from_y(y, measure_markers, px_per_beat)
        extra = {
            "raw_stroke": stroke,
            "shape": "polyline",
            "points": points,
        }
        events.append(NoteEvent(time_beats=time_beats, lane=lane_idx, kind=kind, extra=extra))

    events.sort(key=lambda e: (e.time_beats, e.lane))
    return events

# ----------------------------------------------------------------------
# SectionMetrics aggregation (with rest/hold/overlap & notes_during_hold)
# ----------------------------------------------------------------------

def build_section_metrics(
    notes: List[NoteEvent],
    bpm: float,
    section_beats: float = 16.0,
) -> List[SectionMetrics]:
    """Aggregate NoteEvent list into SectionMetrics instances.

    Computes:
      - npb / nps per section
      - chart avg_npb_chart / avg_nps_chart
      - peak_npb_chart / peak_nps_chart
      - rest_ratio                 : bins without any notes
      - hold_coverage              : bins containing hold/slide events
      - overlap_ratio              : bins with 2+ notes
      - notes_during_hold_ratio    : bins where hold+tap/flick coexist
    """
    if not notes or bpm <= 0.0:
        return []

    notes = sorted(notes, key=lambda n: n.time_beats)
    max_beat = max(n.time_beats for n in notes)
    if max_beat <= 0.0:
        return []

    chart_duration_sec = (max_beat / bpm) * 60.0
    total_notes = len(notes)

    avg_npb_chart = total_notes / max_beat
    avg_nps_chart = total_notes / chart_duration_sec

    section_count = int(max_beat // section_beats) + 1
    sections: List[SectionMetrics] = []

    global_peak_npb = 0.0
    global_peak_nps = 0.0

    bin_beats = 0.25  # quarter-beat bins

    for i in range(section_count):
        start_b = i * section_beats
        end_b = (i + 1) * section_beats

        sec_notes = [n for n in notes if start_b <= n.time_beats < end_b]
        sec_note_count = len(sec_notes)

        section_duration_sec = (section_beats / bpm) * 60.0
        nps = sec_note_count / section_duration_sec if section_duration_sec > 0 else 0.0
        npb = sec_note_count / section_beats if section_beats > 0 else 0.0

        bin_count = max(1, int(section_beats / bin_beats))

        bin_has_note = [False] * bin_count
        bin_has_hold = [False] * bin_count
        bin_note_count = [0] * bin_count
        bin_tap_or_flick_count = [0] * bin_count

        for ev in sec_notes:
            rel = ev.time_beats - start_b
            if rel < 0 or rel >= section_beats:
                continue
            bin_idx = int(rel / bin_beats)
            if bin_idx < 0 or bin_idx >= bin_count:
                continue

            bin_has_note[bin_idx] = True
            bin_note_count[bin_idx] += 1

            if ev.kind in ("hold_body_or_start", "hold_path", "critical_hold_path"):
                bin_has_hold[bin_idx] = True

            if ev.kind in ("tap", "critical_tap", "flick", "flick_arrow"):
                bin_tap_or_flick_count[bin_idx] += 1

        rest_bins = sum(1 for has_note in bin_has_note if not has_note)
        hold_bins = sum(1 for has_hold in bin_has_hold if has_hold)
        overlap_bins = sum(1 for c in bin_note_count if c >= 2)
        ndh_bins = sum(
            1
            for h, tf in zip(bin_has_hold, bin_tap_or_flick_count)
            if h and tf >= 1
        )

        rest_ratio = rest_bins / bin_count
        hold_coverage = hold_bins / bin_count
        overlap_ratio = overlap_bins / bin_count
        notes_during_hold_ratio = ndh_bins / bin_count

        if npb > global_peak_npb:
            global_peak_npb = npb
        if nps > global_peak_nps:
            global_peak_nps = nps

        section = SectionMetrics(
            duration_sec=section_duration_sec,
            bpm=bpm,
            npb=npb,
            nps=nps,
            avg_npb_chart=avg_npb_chart,
            avg_nps_chart=avg_nps_chart,
            peak_npb_chart=0.0,
            peak_nps_chart=0.0,
            rest_ratio=rest_ratio,
            hold_coverage=hold_coverage,
            notes_during_hold_ratio=notes_during_hold_ratio,
            slide_cross_lane_rate=0.0,
            trace_flick_count=0,
            flick_density=0.0,
            overlap_ratio=overlap_ratio,
            lane_cross_rate=0.0,
            spacing_variance=0.0,
            bpm_delta_ratio=0.0,
            bpm_shift_count=0,
            chart_stop_count=0,
            fake_end_flag=False,
        )
        sections.append(section)

    for s in sections:
        s.peak_npb_chart = global_peak_npb
        s.peak_nps_chart = global_peak_nps

    return sections


# ----------------------------------------------------------------------
# Pattern tag detection (NoteEvent + SectionMetrics + severities -> tags)
# ----------------------------------------------------------------------

def detect_pattern_tags(
    notes: List[NoteEvent],
    bpm: float,
    sections: List[SectionMetrics],
    severity_fw: Dict[str, Any],
) -> List[str]:
    tags: set[str] = set()
    if not notes or bpm <= 0:
        return []

    notes = sorted(notes, key=lambda n: n.time_beats)
    max_beat = max(n.time_beats for n in notes)
    chart_duration_sec = (max_beat / bpm) * 60.0 if bpm > 0 else 0.0

    tap_like = [n for n in notes if n.kind in ("tap", "critical_tap")]

    if chart_duration_sec >= 150.0:
        tags.add("duration=>02:30")

    tags |= _detect_jump_and_spacing_tags(notes)
    tags |= _detect_trill_and_stream_tags(tap_like)
    tags |= _detect_stair_tags(tap_like)
    tags |= _detect_drumroll_and_microjack_tags(tap_like)
    tags |= _detect_slide_and_trace_tags(notes)
    tags |= _detect_flick_tags(notes)
    tags |= _detect_readability_tags(notes)
    tags |= _detect_rhythm_tags(tap_like)
    tags |= _detect_hand_assignment_tags(notes)
    tags |= _detect_chord_and_multi_key_tags(tap_like)

    tags |= _detect_severity_based_tags(notes, bpm, sections, severity_fw, chart_duration_sec)

    return sorted(tags)


# ----------------------------------------------------------------------
# NoteEvent-based tag helpers (from Tips Generation Guides sheet 2)
# ----------------------------------------------------------------------

def _detect_jump_and_spacing_tags(notes: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    last: Optional[NoteEvent] = None
    tight_hits = 0

    for n in notes:
        if last is not None:
            dt = n.time_beats - last.time_beats
            lane_delta = abs(n.lane - last.lane)
            if 0.0 < dt <= 0.25:
                tight_hits += 1
            if 2 <= lane_delta <= 3 and dt <= 0.5:
                tags.add("jump")
            if lane_delta >= 4 and dt <= 0.75:
                tags.add("wide_jump")
        last = n

    if tight_hits >= 10:
        tags.add("tight_spacing")
    return tags


def _detect_trill_and_stream_tags(tap_like: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    if len(tap_like) < 4:
        return tags

    short_intervals = 0
    last = tap_like[0]
    for n in tap_like[1:]:
        dt = n.time_beats - last.time_beats
        if 0.0 < dt <= 0.25:
            short_intervals += 1
        last = n
    if short_intervals >= 32:
        tags.add("stream")

    def _scan_for_trill(min_length: int = 7) -> Tuple[bool, bool, bool]:
        vertical = False
        alternating = False
        hybrid = False
        i = 0
        while i < len(tap_like) - 1:
            run = [tap_like[i]]
            j = i + 1
            while j < len(tap_like):
                dt = tap_like[j].time_beats - tap_like[j - 1].time_beats
                if dt <= 0.0 or dt > 0.25:
                    break
                run.append(tap_like[j])
                j += 1
            if len(run) >= min_length:
                lanes = [ev.lane for ev in run]
                unique_lanes = sorted(set(lanes))
                if len(unique_lanes) == 2:
                    vertical = True
                else:
                    groups = ["L" if l <= 4 else "R" for l in lanes]
                    if len(set(groups)) == 2 and groups.count("L") > 0 and groups.count("R") > 0:
                        alternating = True
            i = max(i + 1, j)
        return vertical, alternating, hybrid

    vertical, alternating, hybrid = _scan_for_trill()
    if vertical:
        tags.add("trill_vertical")
    if alternating:
        tags.add("trill_alternating")
    if hybrid:
        tags.add("trill_hybrid")
    return tags


def _detect_stair_tags(tap_like: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    if len(tap_like) < 4:
        return tags

    i = 0
    while i < len(tap_like) - 1:
        run = [tap_like[i]]
        directions: List[int] = []
        j = i + 1
        while j < len(tap_like):
            prev = tap_like[j - 1]
            cur = tap_like[j]
            dt = cur.time_beats - prev.time_beats
            lane_delta = cur.lane - prev.lane
            if dt <= 0.0 or dt > 0.5 or abs(lane_delta) != 1:
                break
            directions.append(1 if lane_delta > 0 else -1)
            run.append(cur)
            j += 1

        if len(run) >= 4:
            tags.add("stair_single")
            avg_lane = sum(ev.lane for ev in run) / len(run)
            if avg_lane <= 4:
                tags.add("stairway_left")
            if avg_lane >= 5:
                tags.add("stairway_right")
            if len(set(directions)) > 1 and directions.count(1) >= 2 and directions.count(-1) >= 2:
                tags.add("zig-zag_stair")
            else:
                if len(directions) >= 4 and directions[0] != directions[-1]:
                    tags.add("spiral_stairway")

        i = max(i + 1, j)

    return tags


def _detect_drumroll_and_microjack_tags(tap_like: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    if not tap_like:
        return tags

    window_beats = 2.0
    min_count = 8

    for i, n in enumerate(tap_like):
        start = n.time_beats
        end = start + window_beats
        window_notes = [ev for ev in tap_like[i:] if start <= ev.time_beats < end]
        if len(window_notes) >= min_count:
            lanes = [ev.lane for ev in window_notes]
            if max(lanes) - min(lanes) <= 1:
                tags.add("drumroll")
                break

    return tags


def _detect_slide_and_trace_tags(notes: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    slides = [n for n in notes if n.kind in ("hold_path", "critical_hold_path")]
    taps_and_flicks = [n for n in notes if n.kind in ("tap", "critical_tap", "flick", "flick_arrow")]

    if not slides:
        return tags

    overlap_count = 0
    for s in slides:
        for ev in taps_and_flicks:
            if abs(ev.time_beats - s.time_beats) <= 0.5:
                overlap_count += 1
                break
    if overlap_count >= 5:
        tags.add("notes_within_slide")

    # low_visibility_trace / zig-zag_slide can be added later with more slide geometry
    return tags


def _detect_flick_tags(notes: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    flicks = [n for n in notes if n.kind in ("flick", "flick_arrow")]
    if not flicks:
        return tags

    flicks.sort(key=lambda n: n.time_beats)

    run_len = 1
    for i in range(1, len(flicks)):
        dt = flicks[i].time_beats - flicks[i - 1].time_beats
        if 0.0 < dt <= 0.25:
            run_len += 1
            if run_len >= 4:
                tags.add("consecutive_flicks")
        else:
            run_len = 1

    slides = [n for n in notes if n.kind in ("hold_path", "critical_hold_path")]
    if slides:
        for f in flicks:
            for s in slides:
                if abs(f.time_beats - s.time_beats) <= 0.5 and abs(f.lane - s.lane) <= 1:
                    tags.add("trace_flick")
                    break
            if "trace_flick" in tags:
                break

    return tags


def _detect_readability_tags(notes: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    bucket_size = 0.0625
    buckets: Dict[int, List[NoteEvent]] = {}

    for n in notes:
        b = int(round(n.time_beats / bucket_size))
        buckets.setdefault(b, []).append(n)

    stacked_hits = 0
    low_vis_hits = 0
    long_short_hits = 0
    tiny_note_hits = 0
    tiny_hold_hits = 0

    for evs in buckets.values():
        if len(evs) >= 3:
            stacked_hits += 1
        widths = [e.extra.get("width_lanes", 1) for e in evs]
        if any(w > 1 for w in widths) and len(evs) >= 2:
            low_vis_hits += 1
        kinds = {e.kind for e in evs}
        if (
            (("hold_body_or_start" in kinds) or ("hold_path" in kinds) or ("critical_hold_path" in kinds))
            and (("tap" in kinds) or ("critical_tap" in kinds))
        ):
            long_short_hits += 1
        for e in evs:
            if e.kind in ("tap", "critical_tap") and e.extra.get("width_lanes", 1) == 1:
                tiny_note_hits += 1
            if e.kind in ("hold_body_or_start", "hold_path", "critical_hold_path") and e.extra.get("width_lanes", 1) == 1:
                tiny_hold_hits += 1

    if stacked_hits >= 3:
        tags.add("stacked_chords")
    if low_vis_hits >= 3:
        tags.add("low_visibility")
    if long_short_hits >= 3:
        tags.add("long_short_taps_mix")
    if tiny_note_hits >= 10:
        tags.add("tiny_notes")
    if tiny_hold_hits >= 5:
        tags.add("tiny_hold")

    return tags


def _detect_rhythm_tags(tap_like: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    if len(tap_like) < 4:
        return tags

    intervals: List[float] = []
    phases: List[float] = []
    last = tap_like[0]
    for n in tap_like[1:]:
        dt = n.time_beats - last.time_beats
        if dt > 0:
            intervals.append(dt)
        last = n

    simple_fracs = [1.0, 0.5, 1.0 / 3.0, 0.25, 1.0 / 6.0]
    off_grid = 0
    for dt in intervals:
        if min(abs(dt - f) for f in simple_fracs) > 0.05:
            off_grid += 1
    if intervals and off_grid / len(intervals) >= 0.3:
        tags.add("difficult_rhythm")

    for n in tap_like:
        phase = n.time_beats - int(n.time_beats)
        phases.append(phase)

    on_beat_count = 0
    off_beat_count = 0
    swing_pair_hits = 0

    for dt in intervals:
        if abs(dt - 1.0 / 3.0) < 0.04 or abs(dt - 2.0 / 3.0) < 0.04:
            swing_pair_hits += 1

    for p in phases:
        if any(abs(p - v) < 0.05 for v in (0.0, 0.25, 0.5, 0.75)):
            on_beat_count += 1
        else:
            off_beat_count += 1

    if swing_pair_hits >= 4:
        tags.add("swing_rhythm")
    if off_beat_count > on_beat_count:
        tags.add("syncopated")

    return tags


def _detect_hand_assignment_tags(notes: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    holds = [n for n in notes if n.kind in ("hold_body_or_start", "hold_path", "critical_hold_path")]
    taps = [n for n in notes if n.kind in ("tap", "critical_tap", "flick", "flick_arrow")]

    if not holds or not taps:
        return tags

    for h in holds:
        for t in taps:
            if abs(t.time_beats - h.time_beats) <= 0.5:
                if (h.lane <= 3 and t.lane >= 7) or (h.lane >= 7 and t.lane <= 3):
                    tags.add("cross_hand")
                    break
        if "cross_hand" in tags:
            break

    left_holds = any(h.lane <= 2 for h in holds)
    right_holds = any(h.lane >= 7 for h in holds)
    if left_holds and right_holds:
        tags.add("forced_hand_swap")

    return tags


def _detect_chord_and_multi_key_tags(tap_like: List[NoteEvent]) -> set[str]:
    tags: set[str] = set()
    if not tap_like:
        return tags

    bucket_size = 0.02
    buckets: Dict[int, List[NoteEvent]] = {}
    for n in tap_like:
        b = int(round(n.time_beats / bucket_size))
        buckets.setdefault(b, []).append(n)

    multi_key_hits = 0
    for evs in buckets.values():
        lanes = {e.lane for e in evs}
        if len(lanes) >= 5:
            multi_key_hits += 1

    if multi_key_hits >= 1:
        tags.add("multi_keys")

    return tags

# ----------------------------------------------------------------------
# Severity + SectionMetrics-based tags
# ----------------------------------------------------------------------

def _detect_severity_based_tags(
    notes: List[NoteEvent],
    bpm: float,
    sections: List[SectionMetrics],
    severity_fw: Dict[str, Any],
    chart_duration_sec: float,
) -> set[str]:
    tags: set[str] = set()

    aggregated: Dict[str, str] = severity_fw.get("aggregated", {}) or {}
    per_section: List[Dict[str, str]] = severity_fw.get("per_section", []) or []

    # low_bpm_high_density
    if "low_bpm_high_density" in aggregated:
        sev = aggregated["low_bpm_high_density"]
        if severity_ge(sev, "mild"):
            tags.add("low_bpm_high_density")

    # bpm_shift
    if "bpm_shift" in aggregated:
        sev = aggregated["bpm_shift"]
        if severity_ge(sev, "light"):
            tags.add("bpm_shift")

    # sudden_speedup / sudden_slowdown via bpm_delta_ratio
    speedup_sections = sum(1 for m in sections if m.bpm_delta_ratio > 0.15)
    slowdown_sections = sum(1 for m in sections if m.bpm_delta_ratio < -0.15)
    if speedup_sections > 0:
        tags.add("sudden_speedup")
    if slowdown_sections > 0:
        tags.add("sudden_slowdown")

    # chart_stop & fake_end
    if "temporal_disruption" in aggregated:
        sev = aggregated["temporal_disruption"]
        if severity_ge(sev, "light"):
            tags.add("chart_stop")
    if any(m.chart_stop_count > 0 for m in sections):
        tags.add("chart_stop")
    if any(m.fake_end_flag for m in sections):
        tags.add("fake_end")

    # stream -> tag "stream"
    stream_sev = aggregated.get("stream")
    if stream_sev and severity_ge(stream_sev, "mild"):
        tags.add("stream")

    # burst -> burst / burst.start / burst.end / post_climax_spike
    burst_sections: List[int] = []
    for idx, sev_map in enumerate(per_section):
        sev = sev_map.get("burst")
        if sev and severity_ge(sev, "mild"):
            burst_sections.append(idx)

    if burst_sections:
        tags.add("burst")
        n_sections = len(sections)
        if n_sections > 0:
            opening_end = max(0, int(n_sections * 0.2))
            ending_start = max(0, int(n_sections * 0.8))
            if any(idx <= opening_end for idx in burst_sections):
                tags.add("burst.start")
            if any(idx >= ending_start for idx in burst_sections):
                tags.add("burst.end")
            tail_start = int(n_sections * 0.7)
            if any(idx >= tail_start for idx in burst_sections):
                tags.add("post_climax_spike")

    # hold_interference -> long_short_taps_mix / notes_within_slide
    hold_int = aggregated.get("hold_interference")
    if hold_int:
        if severity_ge(hold_int, "mild"):
            tags.add("long_short_taps_mix")
        if severity_ge(hold_int, "moderate"):
            tags.add("notes_within_slide")

    # readability -> low_visibility / stacked_chords
    read_sev = aggregated.get("readability")
    if read_sev:
        if severity_ge(read_sev, "light"):
            tags.add("low_visibility")
        if severity_ge(read_sev, "mild"):
            tags.add("stacked_chords")

    return tags
