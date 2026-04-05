"""
Excel Writer for Song Database (full).xlsx
"""

from openpyxl import load_workbook

SHEET_NAMES = {
    "arcaea": "Arcaea",
    "proseka": "Project SEKAI",
    "bandori": "BanG Dream",
}


class ExcelWriter:
    def __init__(self, db_path):
        self.db_path = db_path
        self.wb = load_workbook(db_path)

    def insert_row(self, game_id, canonical_row):
        sheet = self.wb[SHEET_NAMES[game_id]]

        # Prevent duplicates
        song_id = canonical_row["song_id"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0] == song_id:
                return  # Skip duplicate

        ordered = [
            canonical_row.get("song_id"),
            canonical_row.get("name"),
            canonical_row.get("tier"),
            canonical_row.get("level"),
            canonical_row.get("difficulty_label"),
            canonical_row.get("difficulty_code"),
            canonical_row.get("rating_raw"),
            canonical_row.get("note_total_db"),
            canonical_row.get("note_total_chart"),
            canonical_row.get("note_delta"),
            canonical_row.get("chart_path"),
            canonical_row.get("duration_ms"),
            canonical_row.get("bpm"),
        ]
        sheet.append(ordered)

    def save(self):
        self.wb.save(self.db_path)
